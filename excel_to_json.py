import sys
import os
import time
import random
import openpyxl
from openpyxl.styles import  PatternFill

# json转excel - 新建/覆盖
# data = [[1, 2, 3], [4, 5, 6]]
# rules = [[0, 0, 0, 1], [1, 0, 1, 1]]，单元格合并规则[start_row, start_column, end_row, end_column]
def json_to_excel_new(data, filename=None, sheet_name=None, rules=None):
    if not filename:
        filename = '{}{}{}_{}.xlsx'.format(sys.path[0], '\\', time.strftime("%Y%m%d", time.localtime()),
                                           ''.join(random.sample('0123456789', 4)))
    elif not filename.endswith('.xlsx'):
        raise Exception('The type of file should be xlsx.')
    book = openpyxl.Workbook()
    sheet = book.active
    if sheet_name:
        sheet.title = sheet_name
    if data:
        for i in range(len(data)):
            row = data[i]
            for j in range(len(row)):
                sheet.cell(i + 1, j + 1, str(row[j]))
    if rules:
        for rule in rules:
            if len(rule) != 4:
                raise Exception('The rule {} is wrong.'.format(rule))
            else:
                sheet.merge_cells(start_row=rule[0] + 1, start_column=rule[1] + 1, end_row=rule[2] + 1,
                                  end_column=rule[3] + 1)
    book.save(filename)

# json转excel - 新建/追加
# data = [[1, 2, 3], [4, 5, 6]]
# rules = [[0, 0, 0, 1], [1, 0, 1, 1]]，单元格合并规则[start_row, start_column, end_row, end_column]
def json_to_excel_append(data, filename, sheet_name=None, rules=None):
    if not filename:
        raise Exception('The filename can not be empty.')
    if not filename.endswith('.xlsx'):
        raise Exception('The type of file should be xlsx.')
    if not os.path.isfile(filename):
        json_to_excel_new(data, filename, sheet_name, rules)
    else:
        book = openpyxl.load_workbook(filename)
        if sheet_name:
            if book.__contains__(sheet_name):
                sheet = book[sheet_name]
            else:
                sheet = book.create_sheet(sheet_name)
        else:
            sheet = book.active
        # 统计现有行数
        rs = (r for r in sheet.rows)
        cr = sum(1 for _ in rs)
        if data:
            for i in range(len(data)):
                row = data[i]
                for j in range(len(row)):
                    sheet.cell(cr + i + 1, j + 1, str(row[j]))
        if rules:
            for rule in rules:
                if len(rule) != 4:
                    raise Exception('The rule {} is wrong.'.format(rule))
                else:
                    sheet.merge_cells(start_row=cr + rule[0] + 1, start_column=rule[1] + 1, end_row=cr + rule[2] + 1,
                                      end_column=rule[3] + 1)
        book.save(filename)

# excel转json
# return [[1, 2, 3], [4, 5, 6]]
def excel_to_json(filename, sheet_name=None):
    if not filename:
        raise Exception('The filename can not be empty.')
    if not filename.endswith('.xlsx'):
        raise Exception('The type of file should be xlsx.')
    book = openpyxl.load_workbook(filename)
    if sheet_name:
        if not book.__contains__(sheet_name):
            raise Exception('The sheet {} is not exist.'.format(sheet_name))
        else:
            sheet = book[sheet_name]
            # print(sheet)
    else:
        sheet = book.active
        # print(sheet)
    data = []
    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            if cell.value:
                row_data.append(cell.value)
            else:
                row_data.append('')
        data.append(row_data)
    return data

# 填充单元格
# cells = [[0, 2], [1, 2]]
def fill_pass_cell(filename, cells,clear_cells,sheet_name=None, pass_color='00DB00',clear_color='FFFFFF'):
    if cells:
        if not filename:
            raise Exception('The filename can not be empty.')
        if not filename.endswith('.xlsx'):
            raise Exception('The type of file should be xlsx.')
        book = openpyxl.load_workbook(filename)
        if sheet_name:
            if not book.__contains__(sheet_name):
                raise Exception('The sheet {} is not exist.'.format(sheet_name))
            else:
                sheet = book[sheet_name]
        else:
            sheet = book.active
        pass_fill = PatternFill('solid', fgColor=pass_color)
        clear_fill = PatternFill('solid', fgColor=clear_color)

        for cell in cells:
            sheet.cell(cell[0] + 1, cell[1] + 1).fill = pass_fill
            sheet['A'+ str(cell[0] + 1)] = 'pass'

        for cell in clear_cells:
            sheet.cell(cell[0] + 1, cell[1] + 1).fill = clear_fill
            sheet['I' + str(cell[0] + 1)] = ''

        book.save(filename)

def fill_fail_cell(filename, cells, result= 'fail', sheet_name=None,color='FF0000'):
    if cells:
        if not filename:
            raise Exception('The filename can not be empty.')
        if not filename.endswith('.xlsx'):
            raise Exception('The type of file should be xlsx.')
        book = openpyxl.load_workbook(filename)
        if sheet_name:
            if not book.__contains__(sheet_name):
                raise Exception('The sheet {} is not exist.'.format(sheet_name))
            else:
                sheet = book[sheet_name]
        else:
            sheet = book.active
        fill = PatternFill('solid', fgColor=color)
        for cell in cells:
            sheet.cell(cell[0] + 1, cell[1] + 1).fill = fill
            sheet['I'+ str(cell[0] + 1)] = result
            sheet['A' + str(cell[0] + 1)] = 'fail'
        book.save(filename)

def fill_cell(filename, cells, sheet_name=None,color='FFA500'):
    if cells:
        if not filename:
            raise Exception('The filename can not be empty.')
        if not filename.endswith('.xlsx'):
            raise Exception('The type of file should be xlsx.')
        book = openpyxl.load_workbook(filename)
        if sheet_name:
            if not book.__contains__(sheet_name):
                raise Exception('The sheet {} is not exist.'.format(sheet_name))
            else:
                sheet = book[sheet_name]
        else:
            sheet = book.active
        fill = PatternFill('solid', fgColor=color)
        for cell in cells:
            sheet.cell(cell[0] + 1, cell[1] + 1).fill = fill
        book.save(filename)
        # print('done')

if __name__ == '__main__':
    pass
    # data = [['用例名称', '前提条件', '测试步骤', '预期结果'], ['case1', '1、aaa\n2、bbb', '1、步骤1\n2、步骤2\n3、步骤3', '2、结果\n3、果然']]
    # json_to_excel_new(data, 'test4.xlsx')
    # sheet1 = excel_to_json(r'E:\python\tools\datas\test4.xlsx','Sheet1')
    # sheet2 = excel_to_json(r'E:\python\tools\datas\test4.xlsx', 'Sheet2')
    # json_to_excel_append(sheet1, 'test4444444.xlsx','Sheet1',)
    # json_to_excel_append(sheet2, 'test4444444.xlsx', 'Sheet2', )
    # fill_fail_cell('E:\\test_python\\tools\\datas\\test2.xlsx',[[2,8]],result)
    # fill_pass_cell('E:\\test_python\\tools\\datas\\test2.xlsx',[[2,0]],[[2,8]])






