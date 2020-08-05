from openpyxl.styles import Font
import openpyxl
from openpyxl.styles import Alignment
import os, sys


# 处理excel格式—测试用例
def format_excel(file):
    workbook = openpyxl.load_workbook(filename=file)
    sheets = workbook.worksheets

    for sheet in sheets:
        skeys = ['A', 'B']
        lkeys = ['C', 'D', 'E', 'F', 'H', 'J', 'K']
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        font = Font(name='黑体', size=10, bold=False, italic=False, color='000000')
        for s in skeys:
            sheet.column_dimensions[s].width = 19
        for l in lkeys:
            sheet.column_dimensions[l].width = 29

        for index in range(0, sheet.max_row):
            for cell in list(sheet.rows)[index]:
                cell.font = font
                cell.alignment = align
            index = index + 1
            sheet.row_dimensions[index].height = 40

    workbook.save(file)


# 处理excel格式—普通
def format_normal_excel(file):
    workbook = openpyxl.load_workbook(filename=file)
    sheets = workbook.worksheets

    for sheet in sheets:
        skeys = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        font = Font(name='黑体', size=10, bold=False, italic=False, color='000000')
        for s in skeys:
            sheet.column_dimensions[s].width = 10

        for index in range(-1, sheet.max_row):
            for cell in list(sheet.rows)[index]:
                cell.font = font
                cell.alignment = align

            index = index + 1
            sheet.row_dimensions[index].height = 15
    workbook.save(file)


# 删除非第一行
def del_excel(file):
    workbook = openpyxl.load_workbook(filename=file)
    sheets = workbook.worksheets
    ws = sheets[0]
    if ws.max_row > 0:
        ws.delete_rows(2, ws.max_row)
    workbook.save(file)


# 删除全部行
def del_all_excel(file):
    workbook = openpyxl.load_workbook(filename=file)
    sheets = workbook.worksheets
    ws = sheets[0]
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)
    workbook.save(file)

def get_sheets(file):
    workbook = openpyxl.load_workbook(filename=file)
    names = workbook.sheetnames
    return names


# 删除null行的数据
def del_null_row(file):
    workbook = openpyxl.load_workbook(filename=file)
    sheets = workbook.worksheets
    ws = sheets[0]
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row, col).value
            if value == 'null' or value == ' ':
                # print(row,col, value)
                ws.delete_rows(row)
    workbook.save(file)


if __name__ == '__main__':
    file_name = "E:\\python\\tools\\datas\\test2.xlsx"
    get_sheets(file_name)
    # file_name = sys.argv[1]
    # print(file_name)
    # print(sys.argv[1], sys.argv[2])
    #
    # param = sys.argv[2]
    # if param == 'format_excel':
    #     format_excel(file_name)
    # elif param == 'format_normal_excel':
    #     format_normal_excel(file_name)
    # elif param == 'del_excel':
    #     del_excel(file_name)
    # elif param == 'del_all_excel':
    #     del_all_excel(file_name)
    # elif param == 'sheets':
    #     get_sheets(file_name)
    # else:
    #     del_null_row(file_name)
