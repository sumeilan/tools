from openpyxl.styles import Font
import openpyxl
from openpyxl.styles import Alignment
import os

#格式化excel
def format_excel(file):
    workbook = openpyxl.load_workbook(filename=file)
    sheets = workbook.worksheets

    for sheet in sheets:
        skeys = ['A', 'B']
        lkeys = ['C', 'D', 'E', 'F', 'H', 'J', 'K']
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        font = Font(name='黑体', size=10, bold=False, italic=False, color='000000')
        for s in skeys:
            sheet.column_dimensions[s].width = 19
        for l in lkeys:
            sheet.column_dimensions[l].width = 29

        for index in range(0, sheet.max_row):
            for cell in list(sheet.rows)[index]:
                cell.font = font
                cell.alignment = align
            index = index + 1
            sheet.row_dimensions[index].height = 40

    workbook.save(file)

#删除非第一行数据
def del_excel(file):
    workbook = openpyxl.load_workbook(filename=file)
    sheets = workbook.worksheets
    ws  = sheets[0]
    if ws.max_row > 0:
        ws.delete_rows(2, ws.max_row)
        # for i in range(1,ws.max_row+1):
        #     ws.delete_rows(2,i)
    workbook.save(file)

#删除null行的数据
def del_null_row(file):
    workbook = openpyxl.load_workbook(filename=file)
    sheets = workbook.worksheets
    ws = sheets[0]
    for row in range(1,ws.max_row+1):
        for col in range(1,ws.max_column+1):
            value = ws.cell(row, col).value
            if value == 'null'or value== ' ':
                # print(row,col, value)
                ws.delete_rows(row)
    workbook.save(file)


if __name__ == '__main__':
    root = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
    file_name = root + '\\tools\\datas\\test2.xlsx'
    # format_excel(file_name)
    del_null_row(file_name)
